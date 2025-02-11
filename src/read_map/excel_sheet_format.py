

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter



def format_sheet_overview(sheet):
    # def style_all_bgcolor():
    #     return [
    #         'background: #015254',
    #         'color: #ffffff'
    #     ]
    # def style_header():
    #     return [
    #         'background: #013E40',
    #         'color: #013E40',
    #     ]
    # df.set_table_styles([
    #     {
    #         'selector': 'tbody td',
    #         'props': [
    #             ( 'background-color', '#015254' ),
    #             ( 'color', '#ffffff' ),
    #             ( 'border', '3px solid #fff' ),
    #         ],
    #     },
    #     {
    #         'selector': 'tbody th',
    #         'props': [
    #             ( 'background-color', '#013E40' ),
    #             ( 'color', '#013E40' ),
    #             ( 'border', 'none' ),
    #         ],
    #     },
    # ])
    fill_main = PatternFill(start_color='015254',end_color='015254',fill_type='solid')
    fill_header = PatternFill(start_color='013E40',end_color='013E40',fill_type='solid')
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    font_main = Font(color='ffffff')
    font_main_transparent = Font(color='015254')
    font_header = Font(color='013E40')
    border_main = Border(left=Side(style='thick',color='ffffff'),right=Side(style='thick',color='ffffff'),bottom=Side(style='thick',color='ffffff'),top=Side(style='thick',color='ffffff'),)
    border_noborder = Border()
    # alignment_center = Alignment(horizontal='center',vertical='top')
    # alignment_indent = Alignment(indent=2,vertical='top')
    alignment_col1 = Alignment(indent=2,horizontal='right',vertical='top')
    alignment_col2 = Alignment(indent=2,horizontal='left',vertical='top')
    # df.style.apply(style_header,subset=pd.IndexSlice[:1,])
    # sheet.column_dimensions[[s for s in sheet.columns][0].column_letter].width = 200
    # sheet.column_dimensions[[s for s in sheet.columns][1].column_letter].width = 400    
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 125    
    sheet.row_dimensions[1].height = 65
    # for row in [r for r in sheet.rows][0:1]:
    #     for cell in row:
    #         cell.fill = fill_header
    #         cell.font = font_header
    #         row[0].alignment = alignment_indent
    #         row[1].alignment = alignment_center
    for num, row in enumerate([r for r in sheet.rows][2:]):
        sheet.row_dimensions[num+2].height = 25
        for cell in row:
            cell.fill = fill_main
            cell.font = font_main
            cell.border = border_main
            row[0].alignment = alignment_col1
            row[1].alignment = alignment_col2
    for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=26):
        for cell in row:
            cell.fill = fill_header
            cell.font = font_header
            cell.border = border_noborder
    for row in sheet.iter_rows(min_row=2,max_row=999,min_col=1,max_col=26):
        for cell in row:
            cell.fill = fill_main
            cell.font = font_main
    sheet['A2'].font = font_main_transparent
    sheet['A2'].border = border_noborder
    sheet['B2'].border = border_noborder
    sheet.conditional_formatting.add("B4",FormulaRule(formula=["=($B$4=\"Failed\")"],fill=fill_failed))

def format_sheet_variables(sheet):
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    color_shaded =  Font(color='444444')
    sheet.conditional_formatting.add("$D$2:$D$699999",FormulaRule(formula=['=AND(NOT(ISBLANK($A2)),NOT(ISBLANK($D2)),NOT(ISBLANK($F2)),NOT($G2))'],fill=fill_failed))
    sheet.conditional_formatting.add("$D$2:$D$699999",FormulaRule(formula=['=AND(NOT(ISBLANK($A2)),ISBLANK($D2),NOT(ISBLANK($F2)))'],fill=fill_missing))
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 50
    for row_num_within_data_range_zero_based, row in enumerate([r for r in sheet.rows][1:]):
        row_num_openpyxl = row_num_within_data_range_zero_based + 1
        for cell_index, cell in enumerate(row):
            if cell_index==1:
                cell.font = color_shaded
            elif cell_index==2:
                cell.font = color_shaded

def format_sheet_analysisvalues(sheet):
    fill_categorynames = PatternFill(start_color='A9E0A5',end_color='A9E0A5',fill_type='solid')
    fill_shaded = PatternFill(start_color='e9e9e9',end_color='e9e9e9',fill_type='solid')
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    color_shaded =  Font(color='666666')
    # alignment_center = Alignment(horizontal='center',vertical='top')
    alignment_center_top = Alignment(horizontal='center',vertical='top')
    alignment_center_top_wrap = Alignment(wrap_text=True,horizontal='center',vertical='top')
    # row_height = sheet.sheet_format.defaultRowHeight
    for row_num_within_data_range_zero_based, row in enumerate([r for r in sheet.rows][1:]):
        row_num_openpyxl = row_num_within_data_range_zero_based + 1
        if row_num_within_data_range_zero_based % 4 == 0:
            for cell_index, cell in enumerate(row):
                cell.fill = fill_categorynames
                cell.alignment = alignment_center_top_wrap
        elif row_num_within_data_range_zero_based % 4 == 1:
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.fill = fill_shaded
                    cell.alignment = alignment_center_top
        elif row_num_within_data_range_zero_based % 4 == 2:
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.alignment = alignment_center_top
        elif row_num_within_data_range_zero_based % 4 == 3:
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.font = color_shaded
                    cell.alignment = alignment_center_top
        # sheet.row_dimensions[row_num_openpyxl].height = row_height * 2
    sheet.conditional_formatting.add("$C$2:$ZZ$699999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(NOT(ISBLANK(C2)),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_failed))
    # sheet.conditional_formatting.add("$C$2:$ZZ$699999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),TRUE,FALSE),FALSE),FALSE)'],fill=fill_missing))
    sheet.conditional_formatting.add("$C$2:$ZZ$699999",FormulaRule(formula=['IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_missing))
    for col in range(sheet.min_column,sheet.max_column+1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 14
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 25

def format_sheet_validationissues(sheet):
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 85    

def format_with_stdstyles(sheet):
    fill_header = PatternFill(start_color='000000',end_color='000000',fill_type='solid')
    font_header = Font(color='ffffff')
    border_noborder = Border()
    for col in range(sheet.min_column,sheet.max_column+1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 25
    for row in [r for r in sheet.rows][0:1]:
        for cell_index, cell in enumerate(row):
            cell.fill = fill_header
            cell.font = font_header
            if cell_index==0:
                cell.border = border_noborder
    sheet.freeze_panes = sheet['$B$2']



