


from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter



def format_sheet_analysisvalues(sheet):
    fill_categorynames = PatternFill(start_color='A9E0A5',end_color='A9E0A5',fill_type='solid')
    fill_shaded = PatternFill(start_color='e9e9e9',end_color='e9e9e9',fill_type='solid')
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    color_shaded =  Font(color='666666')
    # alignment_center = Alignment(horizontal='center',vertical='top')
    alignment_center_top = Alignment(horizontal='center',vertical='top')
    alignment_center_top_wrap = Alignment(wrap_text=True,horizontal='center',vertical='top')
    # row_height = sheet.sheet_format.defaultRowHeight
    for row_num_within_data_range_zero_based, row in enumerate([r for r in sheet.rows][1:]):
        row_num_openpyxl = row_num_within_data_range_zero_based + 1
        if row_num_within_data_range_zero_based % 4 == 0:
            # category name
            for cell_index, cell in enumerate(row):
                # the first two cells are category name and short name, and then, column C, is the first category
                # but we color the whole row, including 1st 2 cells
                cell.fill = fill_categorynames
                cell.alignment = alignment_center_top_wrap
        elif row_num_within_data_range_zero_based % 4 == 1:
            # category label
            # categories start at column C, so we skip 1st 2 columns
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.fill = fill_shaded
                    cell.alignment = alignment_center_top
        elif row_num_within_data_range_zero_based % 4 == 2:
            # analysis values
            # categories start at column C, so we skip 1st 2 columns
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.alignment = alignment_center_top
                elif cell_index==1:
                    # actually, I'll put some flag if the field is in exclusions here in that cell in column B
                    cell.font = color_shaded
                    cell.alignment = alignment_center_top
        elif row_num_within_data_range_zero_based % 4 == 3:
            # validation flag
            # categories start at column C, so we skip 1st 2 columns
            for cell_index, cell in enumerate(row):
                if cell_index>1:
                    cell.font = color_shaded
                    cell.alignment = alignment_center_top
        # sheet.row_dimensions[row_num_openpyxl].height = row_height * 2
    sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(NOT(ISBLANK(C2)),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_failed))
    # sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),TRUE,FALSE),FALSE),FALSE)'],fill=fill_missing))
    sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_missing))
    for col in range(sheet.min_column,sheet.max_column+1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 14
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 25



