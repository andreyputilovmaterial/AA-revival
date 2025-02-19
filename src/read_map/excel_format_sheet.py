

from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter



if __name__ == '__main__':
    # run as a program
    import excel_format_sheet_overview
    import excel_format_sheet_variables
    import excel_format_sheet_analysisvalues
    import excel_format_sheet_validation
elif '.' in __name__:
    # package
    from . import excel_format_sheet_overview
    from . import excel_format_sheet_variables
    from . import excel_format_sheet_analysisvalues
    from . import excel_format_sheet_validation
else:
    # included with no parent package
    import excel_format_sheet_overview
    import excel_format_sheet_variables
    import excel_format_sheet_analysisvalues
    import excel_format_sheet_validation



format_sheet_overview = excel_format_sheet_overview.format_sheet_overview
format_sheet_variables = excel_format_sheet_variables.format_sheet_variables
format_sheet_analysisvalues = excel_format_sheet_analysisvalues.format_sheet_analysisvalues
format_sheet_validation = excel_format_sheet_validation.format_sheet_validation



def format_with_stdstyles(sheet):
    fill_header = PatternFill(start_color='000000',end_color='000000',fill_type='solid')
    font_header = Font(color='ffffff')
    border_noborder = Border()
    for col in range(sheet.min_column,sheet.max_column+1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 25
    for row in [r for r in sheet.rows][0:1]:
        for cell_index, cell in enumerate(row):
            cell.fill = fill_header
            cell.font = font_header
            if cell_index==0:
                cell.border = border_noborder
    sheet.freeze_panes = sheet['$B$2']



