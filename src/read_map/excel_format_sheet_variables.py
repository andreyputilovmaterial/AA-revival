


from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule



if __name__ == '__main__':
    # run as a program
    import columns_sheet_variables
elif '.' in __name__:
    # package
    from . import columns_sheet_variables
else:
    # included with no parent package
    import columns_sheet_variables



def format_sheet_variables(sheet):
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    color_shaded =  Font(color='444444')
    substitutes = {
        **columns_sheet_variables.column_letters
    }
    sheet.conditional_formatting.add('${col_shortname}$2:${col_shortname}$999999'.format(**substitutes),FormulaRule(formula=['=AND(NOT(ISBLANK(${col_question}2)),NOT(ISBLANK(${col_shortname}2)),NOT(ISBLANK(${col_include}2)),(${col_validation}2="Failed"))'.format(**substitutes)],fill=fill_failed))
    sheet.conditional_formatting.add('${col_shortname}$2:${col_shortname}$999999'.format(**substitutes),FormulaRule(formula=['=AND(NOT(ISBLANK(${col_question}2)),ISBLANK(${col_shortname}2),NOT(ISBLANK(${col_include}2)))'.format(**substitutes)],fill=fill_missing))
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 50
    for row_num_within_data_range_zero_based, row in enumerate([r for r in sheet.rows][1:]):
        # row_num_openpyxl = row_num_within_data_range_zero_based + 1
        for cell_index, cell in enumerate(row):
            if cell_index==1:
                cell.font = color_shaded
            elif cell_index==2:
                cell.font = color_shaded



