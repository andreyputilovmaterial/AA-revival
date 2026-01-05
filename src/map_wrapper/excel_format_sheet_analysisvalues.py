


from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter



def format_sheet_analysisvalues(sheet):
    fill_categorynames = PatternFill(start_color='A9E0A5',end_color='A9E0A5',fill_type='solid')
    fill_shaded = PatternFill(start_color='e9e9e9',end_color='e9e9e9',fill_type='solid')
    fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    color_shaded =  Font(color='666666')
    # alignment_center = Alignment(horizontal='center',vertical='top')
    alignment_center_top = Alignment(horizontal='center',vertical='top')
    alignment_center_top_wrap = Alignment(wrap_text=True,horizontal='center',vertical='top')
    # row_height = sheet.sheet_format.defaultRowHeight
    # return sheet
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=0):
        row_prefix = row[:2]
        row_data   = row[2:]

        mod = i % 4

        if mod == 0:  # category name
            for cell in row:
                cell.fill = fill_categorynames
                cell.alignment = alignment_center_top_wrap

        elif mod == 1:  # category label
            for cell in row_data:
                cell.fill = fill_shaded
                cell.alignment = alignment_center_top

        elif mod == 2:  # analysis values
            for cell in row_data:
                cell.alignment = alignment_center_top
            row_prefix[1].font = color_shaded
            row_prefix[1].alignment = alignment_center_top

        elif mod == 3:  # validation flag
            for cell in row_data:
                cell.font = color_shaded
                cell.alignment = alignment_center_top
        # sheet.row_dimensions[row_num_openpyxl].height = row_height * 2
    
    sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(NOT(ISBLANK(C2)),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_failed))
    # sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['=IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),TRUE,FALSE),FALSE),FALSE)'],fill=fill_missing))
    sheet.conditional_formatting.add("$C$2:$ZZ$999999",FormulaRule(formula=['IF(ISNUMBER(SEARCH("Analysis Value",$A2)),IF(NOT(ISBLANK(C1)),IF(ISBLANK(C2),IF(ISERROR(C3),TRUE,NOT(C3)),FALSE),FALSE),FALSE)'],fill=fill_missing))
    for col in range(sheet.min_column,sheet.max_column+1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 14
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 25



